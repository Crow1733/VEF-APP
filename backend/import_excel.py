"""Carga el inventario del Excel semanal en una BD limpia, lista para producción.

    python import_excel.py "/ruta/al/35.1-SEMANA del 30 al 5 septiembre.xlsx"

Estructura de cada hoja de inventario:

    fila 7   No | Costo | Venta | Gan. | Producto | (existencia) | Entra. | Venta | …
    fila 8   subtítulo de la categoría
    fila 9+  un producto por fila

Dentro de una hoja hay filas con nombre pero SIN costo ni venta: son cabeceras
de sub-grupo, no productos. En la hoja de consignación general esas cabeceras
son además el nombre del consignador, y los productos que siguen son suyos
hasta la siguiente cabecera.

No se pierde nada en silencio: al final imprime cuántas filas entraron y el
detalle de cada una que se descartó, con el motivo.
"""
import sys
import unicodedata
from pathlib import Path

import openpyxl

from database import get_conn, init_db

# ── Hoja → (categoría, es_consignación, consignador fijo) ────────────────────
# consignador None en una hoja de consignación = se toma de las cabeceras.
HOJAS = {
    # Propias del bazar
    "Lienzos y espejos":        ("Lienzos y Espejos",       False, None),
    "Aseo":                     ("Aseo",                    False, None),
    "Ceramica y Reloj":         ("Cerámica y Reloj",         False, None),
    "COSITAS ":                 ("Cositas",                  False, None),
    "TALABARTERIA ":            ("Talabartería",             False, None),
    "Madera-Utiles del hogar":  ("Madera-Útiles del hogar",  False, None),
    "COCINA ":                  ("Cocina",                   False, None),
    "ARREGLOS FLORALES,BICHOS": ("Arreglos y Bisutería",     False, None),
    "Electrodomestico Bazar":   ("Electrodomésticos Bazar",  False, None),
    "Stan Especial":            ("Stan Especial",            False, None),
    "cuentas xpagar":           ("Cuentas por Pagar",        False, None),
    # Consignación
    "Jesus otros":              ("Jesús Bisutería",          True,  "Jesús"),
    "Jesus ropa calzado":       ("Jesús Ropa y Calzado",     True,  "Jesús"),
    "Jesus electrodomesticos":  ("Jesús Electrodomésticos",  True,  "Jesús"),
    "Enrrique":                 ("Enrique Consignación",     True,  "Enrique"),
    "Sucel":                    ("Sucel Cosméticos",         True,  "Sucel"),
    "Cusco-Yanley":             ("Cusco-Yanley Joyería",     True,  "Cusco/Yanley"),
    "CONSIGNACION ":            ("Consignación General",     True,  None),
}

# Cabeceras que son sub-grupos de la propia categoría, no personas. Sin esto se
# tomarían como cambio de consignador en las hojas que sí lo usan.
NO_SON_CONSIGNADOR = {"zapatos yanley", "electrodomesticos", "almacenes cuevita",
                      "cocina jesus"}

FILA_PRIMER_PRODUCTO = 9
COL_NO, COL_COSTO, COL_VENTA, COL_GAN, COL_NOMBRE, COL_STOCK = 0, 1, 2, 3, 4, 5


def num(valor):
    """El Excel mezcla números, textos y celdas con fórmula vacía."""
    if valor is None:
        return None
    if isinstance(valor, (int, float)):
        return float(valor)
    try:
        return float(str(valor).strip().replace(",", "."))
    except ValueError:
        return None


def texto(valor):
    if valor is None:
        return None
    s = " ".join(str(valor).split())   # colapsa saltos y espacios dobles
    return s or None


def sin_tildes(s):
    return "".join(
        c for c in unicodedata.normalize("NFD", s.lower()) if not unicodedata.combining(c)
    )


def buscar_excel() -> Path:
    if len(sys.argv) > 1:
        p = Path(sys.argv[1])
        if not p.exists():
            raise SystemExit(f"No existe el archivo: {p}")
        return p
    raiz = Path(__file__).parent.parent
    candidatos = [
        x for x in list(raiz.glob("*.xlsx")) + list((raiz / "Exel").glob("*.xlsx"))
        if not x.name.startswith("~$")
    ]
    if not candidatos:
        raise SystemExit("Pasa la ruta del .xlsx como argumento")
    return max(candidatos, key=lambda p: p.stat().st_mtime)


def leer_hoja(ws, hoja, consignador_fijo, es_consignacion, descartes):
    """Devuelve los productos de una hoja. Anota en `descartes` lo que deja fuera."""
    productos = []
    consignador_actual = consignador_fijo

    for i, row in enumerate(
        ws.iter_rows(min_row=FILA_PRIMER_PRODUCTO, values_only=True), FILA_PRIMER_PRODUCTO
    ):
        nombre = texto(row[COL_NOMBRE]) if len(row) > COL_NOMBRE else None
        costo = num(row[COL_COSTO]) if len(row) > COL_COSTO else None
        venta = num(row[COL_VENTA]) if len(row) > COL_VENTA else None
        gan = num(row[COL_GAN]) if len(row) > COL_GAN else None
        stock = num(row[COL_STOCK]) if len(row) > COL_STOCK else None

        if not nombre:
            continue                     # fila de relleno (solo el número de orden)

        # Cabecera de sub-grupo: nombre sin precios.
        if not costo and not venta:
            if es_consignacion and sin_tildes(nombre) not in NO_SON_CONSIGNADOR:
                consignador_actual = nombre
                descartes.append((hoja, i, nombre, f"cabecera de consignador → '{nombre}'"))
            else:
                descartes.append((hoja, i, nombre, "cabecera de sub-grupo (sin precios)"))
            continue

        if gan is None and costo is not None and venta is not None:
            gan = venta - costo

        productos.append({
            "fila": i,
            "nombre": nombre,
            "costo": costo or 0.0,
            "precio_venta": venta or 0.0,
            "ganancia": gan or 0.0,
            "stock": stock or 0.0,
            "consignador": consignador_actual if es_consignacion else None,
        })

    return productos


def run():
    ruta = buscar_excel()
    print(f"Excel: {ruta.name}\n")
    init_db()
    wb = openpyxl.load_workbook(str(ruta), data_only=True)

    descartes = []
    por_hoja = {}

    for hoja, (cat, es_cons, consignador) in HOJAS.items():
        if hoja not in wb.sheetnames:
            print(f"  [FALTA] la hoja {hoja!r} no está en el Excel")
            continue
        por_hoja[hoja] = (cat, es_cons, leer_hoja(
            wb[hoja], hoja, consignador, es_cons, descartes
        ))

    # Hojas de inventario del Excel que no están en el mapeo: avisar, no ignorar.
    conocidas = set(HOJAS) | {
        "CUADRE INICIAL ", "CUADRE DE LA SEMANA ", "CUSCO", "Jesus", "Enrrique,",
        "Consignacion.", "Sucel.", "Cuentas por pagar", "cuevita.", "Stan Especial.",
        "Pago de Deudas por Semana ", "Hoja1", "Hoja2", "Hoja3", "Hoja4",
    }
    for h in wb.sheetnames:
        if h not in conocidas:
            print(f"  [AVISO] hoja sin mapear: {h!r}")

    with get_conn() as conn:
        # BD limpia: fuera el catálogo de demo de schema.sql y cualquier operación.
        for t in ("venta_detalle", "ventas", "movimientos_caja", "compra_detalle",
                  "compras", "consignacion_detalle", "consignaciones",
                  "credito_detalle", "pagos_credito", "creditos", "pagos_deuda",
                  "cuentas_por_pagar", "gastos", "bajas", "stock_snapshots",
                  "cierres_semanales", "cajas_config", "cajas", "productos",
                  "categorias"):
            conn.execute(f"DELETE FROM {t}")

        cat_ids = {}
        total = 0
        print(f"\n  {'hoja':28}{'categoría':28}{'prod':>6}{'uds':>8}{'costo':>12}")
        for hoja, (cat, es_cons, prods) in por_hoja.items():
            if cat not in cat_ids:
                cur = conn.execute(
                    "INSERT INTO categorias (nombre, tipo, es_consignacion) VALUES (?,?,?)",
                    (cat, "consignacion" if es_cons else "propia", int(es_cons)),
                )
                cat_ids[cat] = cur.lastrowid

            for p in prods:
                conn.execute(
                    """INSERT INTO productos
                       (categoria_id, nombre, tipo_producto, consignador, costo,
                        precio_venta, ganancia, unidad, stock_inicial, stock_actual,
                        imagen, activa)
                       VALUES (?,?,?,?,?,?,?,?,?,?,?,?)""",
                    (cat_ids[cat], p["nombre"],
                     "consignacion" if es_cons else "propio", p["consignador"],
                     p["costo"], p["precio_venta"], p["ganancia"], "unidad",
                     p["stock"], p["stock"], "", 1),
                )
            total += len(prods)
            uds = sum(p["stock"] for p in prods)
            val = sum(p["stock"] * p["costo"] for p in prods)
            print(f"  {hoja:28}{cat:28}{len(prods):>6}{uds:>8,.0f}{val:>12,.0f}")

        # Las 3 cajas ven todas las categorías.
        for caja in (1, 2, 3):
            for cid in cat_ids.values():
                conn.execute(
                    "INSERT OR IGNORE INTO cajas_config (caja_id, categoria_id) VALUES (?,?)",
                    (caja, cid),
                )

    print(f"\n  {len(cat_ids)} categorías · {total} productos")

    if descartes:
        print(f"\n  ── {len(descartes)} filas NO importadas ──")
        for hoja, fila, nombre, motivo in descartes:
            print(f"     {hoja:26} f{fila:<4} {nombre[:28]:30} {motivo}")


if __name__ == "__main__":
    run()
