from typing import Optional
from io import BytesIO
from fastapi import APIRouter
from fastapi.responses import StreamingResponse
from database import get_conn

router = APIRouter(prefix="/api/reportes", tags=["reportes"])


def _where_ventas(caja_id, desde, hasta) -> tuple[str, list]:
    conds = ["v.estado != 'cancelada'"]
    params = []
    if caja_id is not None:
        conds.append("v.caja_id=?")
        params.append(caja_id)
    if desde:
        conds.append("v.fecha>=?")
        params.append(desde)
    if hasta:
        conds.append("v.fecha<=?")
        params.append(hasta)
    return " AND ".join(conds), params


def _where_movs(caja_id, desde, hasta) -> tuple[str, list]:
    conds = ["1=1"]
    params = []
    if caja_id is not None:
        conds.append("caja_id=?")
        params.append(caja_id)
    if desde:
        conds.append("fecha>=?")
        params.append(desde)
    if hasta:
        conds.append("fecha<=?")
        params.append(hasta)
    return " AND ".join(conds), params


@router.get("/semanal")
def semanal(desde: Optional[str] = None, hasta: Optional[str] = None,
            caja_id: Optional[int] = None):
    with get_conn() as conn:
        cond_v, params_v = _where_ventas(caja_id, desde, hasta)
        cond_m, params_m = _where_movs(caja_id, desde, hasta)

        venta_total = conn.execute(
            f"SELECT COALESCE(SUM(total),0) FROM ventas v WHERE {cond_v} AND es_consignacion=0",
            params_v,
        ).fetchone()[0]
        efectivo_total = conn.execute(
            f"SELECT COALESCE(SUM(subtotal_efectivo),0) FROM ventas v WHERE {cond_v} AND es_consignacion=0",
            params_v,
        ).fetchone()[0]
        transferencia_total = conn.execute(
            f"SELECT COALESCE(SUM(subtotal_transferencia),0) FROM ventas v WHERE {cond_v} AND es_consignacion=0",
            params_v,
        ).fetchone()[0]
        consignacion_total = conn.execute(
            f"SELECT COALESCE(SUM(total),0) FROM ventas v WHERE {cond_v} AND es_consignacion=1",
            params_v,
        ).fetchone()[0]
        utilidad_total = conn.execute(
            f"""SELECT COALESCE(SUM(vd.ganancia_total),0)
                FROM venta_detalle vd
                JOIN ventas v ON vd.venta_id=v.id
                WHERE {cond_v} AND vd.es_consignacion=0""",
            params_v,
        ).fetchone()[0]
        extracciones_total = conn.execute(
            f"SELECT COALESCE(SUM(monto),0) FROM movimientos_caja WHERE {cond_m} AND es_extraccion=1",
            params_m,
        ).fetchone()[0]
        compras_total = conn.execute(
            f"SELECT COALESCE(SUM(monto),0) FROM movimientos_caja WHERE {cond_m} AND es_compra_mercancia=1",
            params_m,
        ).fetchone()[0]

    return {
        "desde": desde, "hasta": hasta, "caja_id": caja_id,
        "venta_total": venta_total, "efectivo_total": efectivo_total,
        "transferencia_total": transferencia_total, "consignacion_total": consignacion_total,
        "utilidad_total": utilidad_total, "extracciones_total": extracciones_total,
        "compras_total": compras_total, "diferencia_caja": 0,
    }


@router.get("/por-categoria")
def por_categoria(desde: Optional[str] = None, hasta: Optional[str] = None,
                  caja_id: Optional[int] = None):
    cond_v, params_v = _where_ventas(caja_id, desde, hasta)
    query = f"""
        SELECT c.id AS categoria_id, c.nombre, c.es_consignacion,
               COALESCE(SUM(vd.subtotal),0) AS venta_total,
               COALESCE(SUM(vd.ganancia_total),0) AS utilidad_total,
               COALESCE(SUM(vd.cantidad),0) AS unidades
        FROM venta_detalle vd
        JOIN ventas v ON vd.venta_id=v.id
        JOIN productos p ON vd.producto_id=p.id
        JOIN categorias c ON p.categoria_id=c.id
        WHERE {cond_v}
        GROUP BY c.id
        ORDER BY venta_total DESC
    """
    with get_conn() as conn:
        rows = conn.execute(query, params_v).fetchall()
    return [dict(r) for r in rows]


@router.get("/consignaciones")
def consignaciones(desde: Optional[str] = None, hasta: Optional[str] = None,
                   caja_id: Optional[int] = None):
    """Liquidación por consignador (equivale a 'IMPORTE AL COSTO' de los ledgers
    Jesus/CUSCO/Sucel…): por cada consignador, lo vendido de sus productos en el
    rango y cuánto se le debe pagar = Σ(cantidad × costo).
    """
    cond_v, params_v = _where_ventas(caja_id, desde, hasta)
    query = f"""
        SELECT COALESCE(NULLIF(TRIM(p.consignador), ''), 'Sin consignador') AS consignador,
               COALESCE(SUM(vd.cantidad), 0)                     AS unidades,
               COALESCE(SUM(vd.cantidad * vd.costo_unitario), 0)  AS a_pagar,
               COALESCE(SUM(vd.subtotal), 0)                      AS venta
        FROM venta_detalle vd
        JOIN ventas v ON vd.venta_id = v.id
        JOIN productos p ON vd.producto_id = p.id
        WHERE {cond_v} AND p.tipo_producto = 'consignacion'
        GROUP BY consignador
        ORDER BY a_pagar DESC
    """
    with get_conn() as conn:
        rows = conn.execute(query, params_v).fetchall()

    consignadores = []
    total = {"unidades": 0, "a_pagar": 0.0, "venta": 0.0, "utilidad_bazar": 0.0}
    for r in rows:
        d = dict(r)
        d["utilidad_bazar"] = d["venta"] - d["a_pagar"]
        consignadores.append(d)
        total["unidades"] += d["unidades"]
        total["a_pagar"] += d["a_pagar"]
        total["venta"] += d["venta"]
        total["utilidad_bazar"] += d["utilidad_bazar"]
    return {"consignadores": consignadores, "total": total}


def compute_cuadre(desde: Optional[str] = None, hasta: Optional[str] = None,
                   socios: int = 2, reserva_pct: float = 20.0, perdida_ganancia: float = 0.0):
    """Cuadre / cierre semanal (CUADRE DE LA SEMANA). P&L global del bazar en el
    rango: ventas − pérdida − costo = utilidad bruta; − gastos − ONAT/arriendo −
    contador − estimulación = utilidad neta; reserva % → dividendos → reparto.
    Es global (todas las cajas), como en el Excel."""
    cond_v, pv = _where_ventas(None, desde, hasta)
    cond_m, pm = _where_movs(None, desde, hasta)

    def _date_clause(col: str):
        c, p = "", []
        if desde:
            c += f" AND {col}>=?"
            p.append(desde)
        if hasta:
            c += f" AND {col}<=?"
            p.append(hasta)
        return c, p

    with get_conn() as conn:
        def sv(expr):
            return conn.execute(f"SELECT COALESCE({expr},0) FROM ventas v WHERE {cond_v}", pv).fetchone()[0]

        venta_total = sv("SUM(total)")
        efectivo = sv("SUM(subtotal_efectivo)")
        transferencia = sv("SUM(subtotal_transferencia)")
        venta_costo = conn.execute(
            f"""SELECT COALESCE(SUM(vd.cantidad*vd.costo_unitario),0)
                FROM venta_detalle vd JOIN ventas v ON vd.venta_id=v.id WHERE {cond_v}""",
            pv,
        ).fetchone()[0]
        perdida_detalle = conn.execute(
            f"""SELECT COALESCE(SUM(vd.perdida_ganancia),0)
                FROM venta_detalle vd JOIN ventas v ON vd.venta_id=v.id WHERE {cond_v}""",
            pv,
        ).fetchone()[0]
        consig_a_pagar = conn.execute(
            f"""SELECT COALESCE(SUM(vd.cantidad*vd.costo_unitario),0)
                FROM venta_detalle vd JOIN ventas v ON vd.venta_id=v.id
                JOIN productos p ON vd.producto_id=p.id
                WHERE {cond_v} AND p.tipo_producto='consignacion'""",
            pv,
        ).fetchone()[0]

        # Gastos por tipo (tabla gastos)
        gc, gp = _date_clause("fecha")

        def sg(tipo):
            return conn.execute(
                f"SELECT COALESCE(SUM(monto),0) FROM gastos WHERE tipo=?{gc}", [tipo, *gp]
            ).fetchone()[0]

        g_salarios = sg("salario")
        g_transporte = sg("transporte")
        g_onat = sg("onat")
        g_arrend = sg("arrendamiento")
        g_contador = sg("contador")
        g_estim = sg("estimulacion")
        g_individual = sg("individual")

        # Movimientos de caja
        def sm(extra):
            return conn.execute(
                f"SELECT COALESCE(SUM(monto),0) FROM movimientos_caja WHERE {cond_m} AND {extra}", pm
            ).fetchone()[0]

        extracciones = sm("es_extraccion=1")
        compras_merc = sm("es_compra_mercancia=1")
        pagos_caja = sm("tipo_movimiento='pago' AND es_extraccion=0 AND es_compra_mercancia=0")

        # Deudas pagadas en el rango
        dc, dp = _date_clause("fecha")
        deudas_pagadas = conn.execute(
            f"SELECT COALESCE(SUM(monto),0) FROM pagos_deuda WHERE 1=1{dc}", dp
        ).fetchone()[0]

        # Faltante/sobrante: suma de diferencias de cajas cerradas en el rango
        fc, fp = _date_clause("fecha_cierre")
        faltante_sobrante = conn.execute(
            f"SELECT COALESCE(SUM(diferencia),0) FROM cajas WHERE estado='cerrada'{fc}", fp
        ).fetchone()[0]

        # Ganancia por elevación de precios (sobreprecio sobre precio de lista)
        ganancia_elevacion = conn.execute(
            f"""SELECT COALESCE(SUM(vd.ganancia_elevacion),0)
                FROM venta_detalle vd JOIN ventas v ON vd.venta_id=v.id WHERE {cond_v}""",
            pv,
        ).fetchone()[0]

        # Ventas a libreta/crédito y cobros del rango (base caja: memo, no entra a venta_total)
        venta_libreta = conn.execute(
            f"SELECT COALESCE(SUM(total),0) FROM creditos WHERE 1=1{dc}", dp
        ).fetchone()[0]
        cobros_libreta = conn.execute(
            f"SELECT COALESCE(SUM(monto),0) FROM pagos_credito WHERE 1=1{dc}", dp
        ).fetchone()[0]

        # Bajas / mermas valoradas al costo
        bajas_total = conn.execute(
            f"SELECT COALESCE(SUM(cantidad*costo_unitario),0) FROM bajas WHERE 1=1{dc}", dp
        ).fetchone()[0]

        # Transferencia desglosada por persona/socio ("Transferencia jesus" vs general)
        ts_rows = conn.execute(
            f"""SELECT COALESCE(NULLIF(v.transferencia_socio,''),'general') soc,
                       COALESCE(SUM(v.subtotal_transferencia),0) m
                FROM ventas v WHERE {cond_v} AND v.subtotal_transferencia>0
                GROUP BY COALESCE(NULLIF(v.transferencia_socio,''),'general')""",
            pv,
        ).fetchall()
        transferencia_por_socio = {r[0]: r[1] for r in ts_rows}

    # ── P&L ──────────────────────────────────────────────────────────────────
    # Pérdida de ganancia = descuentos capturados en ventas + ajuste manual opcional.
    perdida_total = perdida_detalle + perdida_ganancia
    venta_real = venta_total - perdida_total
    utilidad_bruta = venta_real - venta_costo
    gastos_operativos = g_salarios + g_transporte
    onat_arrend = g_onat + g_arrend
    utilidad_neta = utilidad_bruta - gastos_operativos - onat_arrend - g_contador - g_estim
    reserva = utilidad_neta * reserva_pct / 100
    dividendos = utilidad_neta - reserva
    por_socio = dividendos / socios if socios else dividendos
    efectivo_caja = efectivo - extracciones - compras_merc - pagos_caja - deudas_pagadas

    return {
        "desde": desde, "hasta": hasta,
        "venta_total": venta_total, "efectivo": efectivo, "transferencia": transferencia,
        "perdida_ganancia": perdida_total, "venta_real": venta_real,
        "venta_costo": venta_costo, "utilidad_bruta": utilidad_bruta,
        "gastos": {
            "salarios": g_salarios, "transporte": g_transporte,
            "onat": g_onat, "arrendamiento": g_arrend, "contador": g_contador,
            "estimulacion": g_estim, "individual": g_individual,
            "operativos": gastos_operativos,
        },
        "utilidad_neta": utilidad_neta,
        "reserva_pct": reserva_pct, "reserva": reserva,
        "dividendos": dividendos, "socios": socios, "por_socio": por_socio,
        "movimientos": {
            "extracciones": extracciones, "compras_mercancia": compras_merc,
            "pagos_caja": pagos_caja, "deudas_pagadas": deudas_pagadas,
        },
        "consignadores_a_pagar": consig_a_pagar,
        "faltante_sobrante": faltante_sobrante,
        "efectivo_caja": efectivo_caja,
        "ganancia_elevacion": ganancia_elevacion,
        "venta_libreta": venta_libreta,
        "cobros_libreta": cobros_libreta,
        "bajas_total": bajas_total,
        "transferencia_por_socio": transferencia_por_socio,
    }


@router.get("/cuadre")
def cuadre(desde: Optional[str] = None, hasta: Optional[str] = None,
           socios: int = 2, reserva_pct: float = 20.0, perdida_ganancia: float = 0.0):
    return compute_cuadre(desde, hasta, socios, reserva_pct, perdida_ganancia)


@router.get("/inventario")
def inventario(caja_id: Optional[int] = None):
    """Inventario valorado por categoría (equivale a 'IMPORTE EN MERCANCIA' del
    CUADRE INICIAL): unidades en stock, valor al costo, valor a la venta y
    utilidad potencial. Si se pasa caja_id, limita a las categorías de esa caja.
    """
    cat_filter = ""
    params: list = []
    if caja_id is not None:
        cat_filter = "AND c.id IN (SELECT categoria_id FROM cajas_config WHERE caja_id=?)"
        params.append(caja_id)
    query = f"""
        SELECT c.id AS categoria_id, c.nombre, c.es_consignacion,
               COUNT(p.id) AS productos,
               COALESCE(SUM(p.stock_actual), 0)                AS unidades,
               COALESCE(SUM(p.stock_actual * p.costo), 0)       AS valor_costo,
               COALESCE(SUM(p.stock_actual * p.precio_venta), 0) AS valor_venta
        FROM categorias c
        LEFT JOIN productos p ON p.categoria_id = c.id AND p.activa = 1
        WHERE c.activa = 1 {cat_filter}
        GROUP BY c.id
        ORDER BY valor_costo DESC
    """
    with get_conn() as conn:
        rows = conn.execute(query, params).fetchall()

    categorias = []
    total = {"productos": 0, "unidades": 0, "valor_costo": 0.0,
             "valor_venta": 0.0, "utilidad_potencial": 0.0}
    for r in rows:
        d = dict(r)
        d["utilidad_potencial"] = d["valor_venta"] - d["valor_costo"]
        categorias.append(d)
        total["productos"] += d["productos"]
        total["unidades"] += d["unidades"]
        total["valor_costo"] += d["valor_costo"]
        total["valor_venta"] += d["valor_venta"]
        total["utilidad_potencial"] += d["utilidad_potencial"]
    return {"categorias": categorias, "total": total}


@router.get("/utilidad-por-producto")
def utilidad_por_producto(desde: Optional[str] = None, hasta: Optional[str] = None,
                          caja_id: Optional[int] = None):
    cond_v, params_v = _where_ventas(caja_id, desde, hasta)
    query = f"""
        SELECT p.id AS producto_id, p.nombre, c.nombre AS categoria_nombre,
               p.tipo_producto,
               CASE WHEN p.tipo_producto='consignacion' THEN 1 ELSE 0 END AS es_consignacion,
               COALESCE(SUM(vd.cantidad),0) AS unidades,
               COALESCE(SUM(vd.subtotal),0) AS venta_total,
               COALESCE(SUM(vd.ganancia_total),0) AS utilidad_total
        FROM venta_detalle vd
        JOIN ventas v ON vd.venta_id=v.id
        JOIN productos p ON vd.producto_id=p.id
        JOIN categorias c ON p.categoria_id=c.id
        WHERE {cond_v}
        GROUP BY p.id
        ORDER BY utilidad_total DESC
    """
    with get_conn() as conn:
        rows = conn.execute(query, params_v).fetchall()
    return [dict(r) for r in rows]


def _dias_semana(desde: str, hasta: str) -> list[str]:
    from datetime import datetime, timedelta
    d = datetime.strptime(desde[:10], "%Y-%m-%d")
    h = datetime.strptime(hasta[:10], "%Y-%m-%d")
    dias = []
    while d <= h:
        dias.append(d.strftime("%Y-%m-%d"))
        d += timedelta(days=1)
    return dias


def _dia_nombre(fecha_iso: str) -> str:
    nombres = {0: "Domingo", 1: "Lunes", 2: "Martes", 3: "Miercoles",
               4: "Jueves", 5: "Viernes", 6: "Sabado"}
    from datetime import datetime
    return nombres[datetime.strptime(fecha_iso[:10], "%Y-%m-%d").weekday()]


def _serie_por_dia(conn, query: str, params) -> dict:
    """Ejecuta una consulta agrupada por fecha (substr 1..10) y devuelve
    {'YYYY-MM-DD': valor} para mapear rapido sobre el rango de dias."""
    return {r[0]: r[1] for r in conn.execute(query, params).fetchall()}


@router.get("/movimientos-diarios")
def movimientos_diarios(desde: Optional[str] = None, hasta: Optional[str] = None,
                        caja_id: Optional[int] = None):
    """Movimientos de inventario diarios por producto: entradas (compras+consignaciones),
    ventas, y stock final por cada dia de la semana, agrupados por categoria."""
    if not desde or not hasta:
        return {"categorias": [], "dias": []}

    dias = _dias_semana(desde, hasta)
    with get_conn() as conn:
        cats = conn.execute(
            "SELECT id, nombre FROM categorias WHERE activa=1 ORDER BY nombre"
        ).fetchall()

        categorias_out = []
        for cat in cats:
            prods = conn.execute(
                "SELECT id, nombre, costo, precio_venta, stock_actual "
                "FROM productos WHERE categoria_id=? AND activa=1 ORDER BY nombre",
                (cat["id"],),
            ).fetchall()

            productos_out = []
            cat_venta_total = 0
            cat_utilidad_total = 0
            cat_stock_final = 0

            for p in prods:
                prod_dias = []
                total_entradas = 0
                total_ventas = 0

                for dia in dias:
                    dia_sig = dia + " 23:59:59"
                    entradas = conn.execute(
                        """SELECT COALESCE(SUM(cd.cantidad),0)
                           FROM compra_detalle cd
                           JOIN compras c ON cd.compra_id=c.id
                           WHERE cd.producto_id=? AND c.fecha>=? AND c.fecha<=?""",
                        (p["id"], dia, dia_sig),
                    ).fetchone()[0]
                    entradas_consig = conn.execute(
                        """SELECT COALESCE(SUM(cd.cantidad_entregada),0)
                           FROM consignacion_detalle cd
                           JOIN consignaciones c ON cd.consignacion_id=c.id
                           WHERE cd.producto_id=? AND c.fecha_inicio>=? AND c.fecha_inicio<=?""",
                        (p["id"], dia, dia_sig),
                    ).fetchone()[0]
                    ventas_dia = conn.execute(
                        """SELECT COALESCE(SUM(vd.cantidad),0)
                           FROM venta_detalle vd
                           JOIN ventas v ON vd.venta_id=v.id
                           WHERE vd.producto_id=? AND v.estado!='cancelada'
                           AND v.fecha>=? AND v.fecha<=?""",
                        (p["id"], dia, dia_sig),
                    ).fetchone()[0]

                    total_entradas += entradas + entradas_consig
                    total_ventas += ventas_dia
                    prod_dias.append({
                        "fecha": dia,
                        "dia": _dia_nombre(dia),
                        "entradas": entradas + entradas_consig,
                        "ventas": ventas_dia,
                    })

                venta_prod = conn.execute(
                    """SELECT COALESCE(SUM(vd.subtotal),0), COALESCE(SUM(vd.ganancia_total),0)
                       FROM venta_detalle vd JOIN ventas v ON vd.venta_id=v.id
                       WHERE vd.producto_id=? AND v.estado!='cancelada'
                       AND v.fecha>=? AND v.fecha<=?""",
                    (p["id"], desde, hasta + " 23:59:59"),
                ).fetchone()

                stock_final = p["stock_actual"]
                cat_venta_total += venta_prod[0]
                cat_utilidad_total += venta_prod[1]
                cat_stock_final += stock_final

                productos_out.append({
                    "producto_id": p["id"],
                    "nombre": p["nombre"],
                    "costo": p["costo"],
                    "precio_venta": p["precio_venta"],
                    "dias": prod_dias,
                    "venta_total": venta_prod[0],
                    "utilidad_total": venta_prod[1],
                    "stock_final": stock_final,
                })

            if productos_out:
                categorias_out.append({
                    "categoria_id": cat["id"],
                    "categoria_nombre": cat["nombre"],
                    "productos": productos_out,
                    "venta_total": cat_venta_total,
                    "utilidad_total": cat_utilidad_total,
                    "stock_final": cat_stock_final,
                })

    return {"categorias": categorias_out, "dias": dias}


@router.get("/cobro-diario")
def cobro_diario(desde: Optional[str] = None, hasta: Optional[str] = None,
                 caja_id: Optional[int] = None):
    """Cobro en efectivo por dia y por categoria (columnas AA-AN del Excel)."""
    if not desde or not hasta:
        return {"dias": [], "categorias": [], "totales_por_dia": {}}

    dias = _dias_semana(desde, hasta)
    cond_v, params_v = _where_ventas(caja_id, desde, hasta + " 23:59:59")

    with get_conn() as conn:
        cats = conn.execute(
            "SELECT id, nombre FROM categorias WHERE activa=1 ORDER BY nombre"
        ).fetchall()

        cat_dias = {}
        for cat in cats:
            cat_dias[cat["id"]] = {"nombre": cat["nombre"], "montos": {}}
            for dia in dias:
                dia_sig = dia + " 23:59:59"
                monto = conn.execute(
                    f"""SELECT COALESCE(SUM(vd.subtotal),0)
                        FROM venta_detalle vd
                        JOIN ventas v ON vd.venta_id=v.id
                        JOIN productos p ON vd.producto_id=p.id
                        WHERE p.categoria_id=? AND v.estado!='cancelada'
                        AND v.fecha>=? AND v.fecha<=?""",
                    (cat["id"], dia, dia_sig),
                ).fetchone()[0]
                cat_dias[cat["id"]]["montos"][dia] = monto

        totales_por_dia = {}
        for dia in dias:
            totales_por_dia[dia] = sum(
                cat_dias[c["id"]]["montos"].get(dia, 0) for c in cats
            )

    return {
        "dias": dias,
        "categorias": [
            {"categoria_id": c["id"], "nombre": cat_dias[c["id"]]["nombre"],
             "montos": cat_dias[c["id"]]["montos"]}
            for c in cats
        ],
        "totales_por_dia": totales_por_dia,
    }


@router.get("/export-excel")
def export_excel(desde: Optional[str] = None, hasta: Optional[str] = None):
    """Genera el Excel semanal con la estructura exacta del template."""
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from datetime import datetime

    if not desde or not hasta:
        desde = datetime.now().strftime("%Y-%m-01")
        hasta = datetime.now().strftime("%Y-%m-%d")

    wb = openpyxl.Workbook()

    header_font = Font(bold=True, size=12)
    title_font = Font(bold=True, size=14)
    money_fmt = '#,##0'
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

    dias = _dias_semana(desde, hasta)
    n = len(dias)
    gl = openpyxl.utils.get_column_letter

    # Etiqueta corta por dia (estilo plantilla): L-11, M-12, Mi-13, J-14...
    letra_dia = {0: "L", 1: "M", 2: "Mi", 3: "J", 4: "V", 5: "S", 6: "D"}
    meses = {1: "Enero", 2: "Febrero", 3: "Marzo", 4: "Abril", 5: "Mayo",
             6: "Junio", 7: "Julio", 8: "Agosto", 9: "Septiembre",
             10: "Octubre", 11: "Noviembre", 12: "Diciembre"}
    d_ini = datetime.strptime(dias[0], "%Y-%m-%d")
    d_fin = datetime.strptime(dias[-1], "%Y-%m-%d")
    titulo_semana = f"Semana del {d_ini.day} al {d_fin.day} de {meses[d_fin.month]}"

    # Posiciones de columna (1-indexed) replicando la plantilla:
    # A No | B Costo | C Venta | D Gan | E Producto | F Existencia inicial |
    # G.. pares Entra/Venta por dia | Venta Total | Deuda | Baja sem |
    # Stock final | Entradas totales | (ocultas) valor entradas, cobro diario,
    # valoraciones (importe entrada/venta/utilidad/stock al costo).
    col_ent = [7 + 2 * i for i in range(n)]   # entradas por dia
    col_ven = [8 + 2 * i for i in range(n)]   # ventas por dia
    c_total = 7 + 2 * n        # U: Venta Total
    c_deuda = c_total + 1      # V: Deuda (unidades a credito/libreta)
    c_baja = c_total + 2       # W: Baja sem. (mermas)
    c_stock = c_total + 3      # X: Stock final
    c_entot = c_total + 4      # Y: Entradas totales
    c_zvalor = c_total + 5     # Z: entradas totales * costo
    c_cobro0 = c_total + 6     # AA..: cobro efectivo por dia
    c_cuadre0 = c_cobro0 + n   # AH..: cuadre en efectivo (total por dia de la categoria)
    c_val0 = max(41, c_cuadre0 + n)   # AO..: bloque de valoraciones

    with get_conn() as conn:
        cats = conn.execute(
            "SELECT id, nombre FROM categorias WHERE activa=1 ORDER BY nombre"
        ).fetchall()

        # Datos por producto para las columnas Deuda (V), Baja (W) y stock inicial (F).
        r_ini, r_fin = dias[0], dias[-1] + " 23:59:59"
        deuda_por_prod = _serie_por_dia(conn,
            """SELECT cd.producto_id, COALESCE(SUM(cd.cantidad),0)
               FROM credito_detalle cd JOIN creditos c ON cd.credito_id=c.id
               WHERE c.fecha>=? AND c.fecha<=? GROUP BY cd.producto_id""", (r_ini, r_fin))
        baja_por_prod = _serie_por_dia(conn,
            """SELECT producto_id, COALESCE(SUM(cantidad),0) FROM bajas
               WHERE fecha>=? AND fecha<=? GROUP BY producto_id""", (r_ini, r_fin))
        # Stock inicial = último snapshot con fecha <= inicio de semana (si existe).
        stock_ini_snap = {}
        for sr in conn.execute(
            "SELECT producto_id, stock FROM stock_snapshots WHERE fecha<=? ORDER BY producto_id, fecha, id",
            (dias[0] + " 23:59:59",),
        ).fetchall():
            stock_ini_snap[sr[0]] = sr[1]   # ordenado asc -> el último gana

        first_sheet = True
        for cat in cats:
            if first_sheet:
                ws = wb.active
                ws.title = cat["nombre"][:31]
                first_sheet = False
            else:
                ws = wb.create_sheet(title=cat["nombre"][:31])

            prods = conn.execute(
                "SELECT id, nombre, costo, precio_venta, stock_inicial, stock_actual "
                "FROM productos WHERE categoria_id=? AND activa=1 ORDER BY nombre",
                (cat["id"],),
            ).fetchall()

            # ── Cabeceras (filas 1, 5, 6, 7, 8) ────────────────────────────
            ws["A1"] = "Basar Ojos Color Sol "
            ws["A1"].font = title_font
            ws.cell(row=1, column=c_entot, value="Entradas totales de la semana")

            ws.cell(row=5, column=1, value="INVENTARIO ").font = header_font
            f5 = ws.cell(row=5, column=6, value=d_ini)
            f5.number_format = "dd/mm/yy"
            ws.cell(row=5, column=7, value=titulo_semana).font = header_font
            ws.cell(row=5, column=c_total, value="Venta Total ").font = Font(bold=True)
            ws.cell(row=5, column=c_deuda, value="Deuda ").font = Font(bold=True)
            ws.cell(row=5, column=c_baja, value="Baja sem.").font = Font(bold=True)
            ws.cell(row=5, column=c_stock,
                    value=f"F. {d_fin.strftime('%d/%m/%y')}").font = Font(bold=True)

            for i, dia in enumerate(dias):
                d = datetime.strptime(dia, "%Y-%m-%d")
                etq = f"{letra_dia[d.weekday()]}-{d.day}"
                ws.cell(row=6, column=col_ent[i], value=etq).font = Font(bold=True)

            enc = {1: "No", 2: "Costo ", 3: "Venta ", 4: "Gan.", 5: "Producto "}
            for c_idx, txt in enc.items():
                cell = ws.cell(row=7, column=c_idx, value=txt)
                cell.font = Font(bold=True)
                cell.fill = header_fill
                cell.border = thin_border
            for i in range(n):
                e = ws.cell(row=7, column=col_ent[i], value="Entra.")
                e.font = Font(bold=True); e.fill = header_fill; e.border = thin_border
                v = ws.cell(row=7, column=col_ven[i], value="Venta")
                v.font = Font(bold=True); v.fill = header_fill; v.border = thin_border
            # Cabecera del bloque "Cuadre en efectivo de la Semana" (totales por dia)
            cab_q = ws.cell(row=7, column=c_cuadre0, value="Cuadre en efectivo de la Semana")
            cab_q.font = Font(bold=True); cab_q.fill = header_fill

            ws.cell(row=8, column=5, value=cat["nombre"]).font = Font(bold=True, italic=True)

            # ── Filas de producto: valores reales + formulas del template ──
            row = 8
            last_row = 8 + len(prods)
            for idx, p in enumerate(prods, 1):
                row += 1
                r = row
                ws.cell(row=r, column=1, value=idx).border = thin_border
                cell_b = ws.cell(row=r, column=2, value=p["costo"])
                cell_b.number_format = money_fmt; cell_b.border = thin_border
                cell_c = ws.cell(row=r, column=3, value=p["precio_venta"])
                cell_c.number_format = money_fmt; cell_c.border = thin_border
                ws.cell(row=r, column=4, value=f"=C{r}-B{r}").border = thin_border
                ws.cell(row=r, column=5, value=p["nombre"]).border = thin_border
                # F: stock inicial = snapshot del cierre anterior si existe, si no el del producto
                ws.cell(row=r, column=6,
                        value=stock_ini_snap.get(p["id"], p["stock_inicial"] or 0)).border = thin_border

                # Entradas (compras + consignaciones) y ventas por dia (valores)
                ini, fin = dias[0], dias[-1] + " 23:59:59"
                ent_dia = _serie_por_dia(conn,
                    """SELECT substr(c.fecha,1,10) d, COALESCE(SUM(cd.cantidad),0)
                       FROM compra_detalle cd JOIN compras c ON cd.compra_id=c.id
                       WHERE cd.producto_id=? AND c.fecha>=? AND c.fecha<=?
                       GROUP BY substr(c.fecha,1,10)""", (p["id"], ini, fin))
                ent_cons = _serie_por_dia(conn,
                    """SELECT substr(c.fecha_inicio,1,10) d, COALESCE(SUM(cd.cantidad_entregada),0)
                       FROM consignacion_detalle cd JOIN consignaciones c ON cd.consignacion_id=c.id
                       WHERE cd.producto_id=? AND c.fecha_inicio>=? AND c.fecha_inicio<=?
                       GROUP BY substr(c.fecha_inicio,1,10)""", (p["id"], ini, fin))
                ven_dia = _serie_por_dia(conn,
                    """SELECT substr(v.fecha,1,10) d, COALESCE(SUM(vd.cantidad),0)
                       FROM venta_detalle vd JOIN ventas v ON vd.venta_id=v.id
                       WHERE vd.producto_id=? AND v.estado!='cancelada'
                       AND v.fecha>=? AND v.fecha<=?
                       GROUP BY substr(v.fecha,1,10)""", (p["id"], ini, fin))

                for i, dia in enumerate(dias):
                    ent = (ent_dia.get(dia, 0) or 0) + (ent_cons.get(dia, 0) or 0)
                    ws.cell(row=r, column=col_ent[i], value=ent).border = thin_border
                    ws.cell(row=r, column=col_ven[i],
                            value=ven_dia.get(dia, 0) or 0).border = thin_border

                # V: Deuda (unidades a credito) y W: Baja sem. (mermas) del rango
                deu = deuda_por_prod.get(p["id"], 0) or 0
                baj = baja_por_prod.get(p["id"], 0) or 0
                if deu:
                    ws.cell(row=r, column=c_deuda, value=deu).border = thin_border
                if baj:
                    ws.cell(row=r, column=c_baja, value=baj).border = thin_border

                ent_cells = [f"{gl(c)}{r}" for c in col_ent]
                ven_cells = [f"{gl(c)}{r}" for c in col_ven]
                # U: Venta Total = suma de las ventas diarias
                ws.cell(row=r, column=c_total, value="=" + "+".join(ven_cells))
                # X: Stock final = inicial + entradas − ventas − deuda − baja
                ws.cell(row=r, column=c_stock,
                        value=(f"=F{r}+" + "+".join(ent_cells) + "-" + "-".join(ven_cells)
                               + f"-{gl(c_deuda)}{r}-{gl(c_baja)}{r}"))
                # Y: Entradas totales = suma de las entradas diarias
                ws.cell(row=r, column=c_entot, value="=" + "+".join(ent_cells))
                # Z: entradas totales valoradas al costo
                zc = ws.cell(row=r, column=c_zvalor, value=f"={gl(c_entot)}{r}*B{r}")
                zc.number_format = money_fmt
                # Cobro efectivo por dia = ventas del dia * precio de venta
                for i in range(n):
                    cc = ws.cell(row=r, column=c_cobro0 + i,
                                 value=f"={gl(col_ven[i])}{r}*C{r}")
                    cc.number_format = money_fmt
                # AH..: Cuadre en efectivo = total de cada dia para la categoria
                for i in range(n):
                    cl = gl(c_cobro0 + i)
                    qc = ws.cell(row=r, column=c_cuadre0 + i,
                                 value=f"=SUM({cl}9:{cl}{last_row})")
                    qc.number_format = money_fmt
                # Valoraciones: AP=Venta*Costo, AQ=Venta*Precio, AR=utilidad, AS=Stock*Costo
                ws.cell(row=r, column=c_val0, value=idx)
                ap = ws.cell(row=r, column=c_val0 + 1, value=f"={gl(c_total)}{r}*B{r}")
                aq = ws.cell(row=r, column=c_val0 + 2, value=f"={gl(c_total)}{r}*C{r}")
                ar = ws.cell(row=r, column=c_val0 + 3,
                             value=f"={gl(c_val0 + 2)}{r}-{gl(c_val0 + 1)}{r}")
                as_ = ws.cell(row=r, column=c_val0 + 4, value=f"={gl(c_stock)}{r}*B{r}")
                for cell in (ap, aq, ar, as_):
                    cell.number_format = money_fmt

            # Anchos de las columnas visibles principales
            ws.column_dimensions["A"].width = 5
            ws.column_dimensions["B"].width = 9
            ws.column_dimensions["C"].width = 9
            ws.column_dimensions["D"].width = 8
            ws.column_dimensions["E"].width = 26

    # ── Hoja CUADRE INICIAL (mercancia / balance inicial de la semana) ───────
    ws_ci = wb.create_sheet(title="CUADRE INICIAL")
    ws_ci["A1"] = "Bazar Ojos Color Sol - Cierre de la semana"
    ws_ci["A1"].font = title_font
    ws_ci["A2"] = titulo_semana
    ci = compute_cuadre(dias[0], dias[-1] + " 23:59:59", 2, 20.0, 0.0)
    with get_conn() as conn:
        venta_dia_ci = _serie_por_dia(conn,
            """SELECT substr(fecha,1,10) d, COALESCE(SUM(total),0)
               FROM ventas WHERE estado!='cancelada' AND fecha>=? AND fecha<=?
               GROUP BY substr(fecha,1,10)""", (dias[0], dias[-1] + " 23:59:59"))
        merc_cat = conn.execute(
            """SELECT c.nombre, COALESCE(SUM(p.stock_actual*p.costo),0) costo
               FROM categorias c LEFT JOIN productos p
                 ON p.categoria_id=c.id AND p.activa=1
               WHERE c.activa=1 GROUP BY c.id ORDER BY costo DESC""").fetchall()
        merc_total = sum(r[1] for r in merc_cat)
        entradas_costo = conn.execute(
            """SELECT COALESCE(SUM(cd.subtotal),0) FROM compra_detalle cd
               JOIN compras c ON cd.compra_id=c.id WHERE c.fecha>=? AND c.fecha<=?""",
            (dias[0], dias[-1] + " 23:59:59")).fetchone()[0]
        entradas_costo += conn.execute(
            """SELECT COALESCE(SUM(cd.cantidad_entregada*cd.costo_acordado),0)
               FROM consignacion_detalle cd JOIN consignaciones c ON cd.consignacion_id=c.id
               WHERE c.fecha_inicio>=? AND c.fecha_inicio<=?""",
            (dias[0], dias[-1] + " 23:59:59")).fetchone()[0]

    # Tabla 1: Cuadre en efectivo de la Semana (venta por dia)
    ws_ci.cell(row=4, column=1, value="Cuadre en efectivo de la Semana").font = header_font
    rci = 5
    for dia in dias:
        ws_ci.cell(row=rci, column=1, value=_dia_nombre(dia))
        ws_ci.cell(row=rci, column=2, value=venta_dia_ci.get(dia, 0) or 0).number_format = money_fmt
        rci += 1
    ws_ci.cell(row=rci, column=1, value="TOTAL SEMANAL").font = Font(bold=True)
    tci = ws_ci.cell(row=rci, column=2, value=ci["venta_total"])
    tci.number_format = money_fmt
    tci.font = Font(bold=True)

    # Tabla 2: Totales financieros
    rci += 2
    ws_ci.cell(row=rci, column=1, value="Totales").font = header_font
    for label, val in [
        ("Total de Venta al Costo", ci["venta_costo"]),
        ("Total de Venta a la Venta", ci["venta_total"]),
        ("Utilidad Total", ci["utilidad_bruta"]),
        ("Importe en Mercancia (al costo)", merc_total),
    ]:
        rci += 1
        ws_ci.cell(row=rci, column=1, value=label)
        ws_ci.cell(row=rci, column=2, value=val).number_format = money_fmt

    # Tabla 3: Mercancia del Bazar por categoria
    rci += 2
    ws_ci.cell(row=rci, column=1, value="Mercancia del Bazar por categoria").font = header_font
    for nombre, costo in merc_cat:
        rci += 1
        ws_ci.cell(row=rci, column=1, value=nombre)
        ws_ci.cell(row=rci, column=2, value=costo).number_format = money_fmt

    # Tabla 4: Importes de inicio de semana
    rci += 2
    ws_ci.cell(row=rci, column=1, value="Importes de inicio de semana").font = header_font
    for label, val in [
        ("Importe en Mercancia (actual, al costo)", merc_total),
        ("Entradas de la Semana (al costo)", entradas_costo),
        ("Total en Mercancia + Entradas", merc_total + entradas_costo),
    ]:
        rci += 1
        ws_ci.cell(row=rci, column=1, value=label)
        ws_ci.cell(row=rci, column=2, value=val).number_format = money_fmt

    ws_ci.column_dimensions["A"].width = 38
    ws_ci.column_dimensions["B"].width = 18

    # ── Hoja CUADRE DE LA SEMANA (desglose diario + P&L) ────────────────────
    ws_c = wb.create_sheet(title="CUADRE DE LA SEMANA")
    ws_c["A1"] = "Basar Ojos Color Sol  -  Cuadre de la Semana"
    ws_c["A1"].font = title_font
    ws_c["A2"] = titulo_semana
    ws_c.cell(row=6, column=6, value="Gastos General ").font = Font(bold=True)
    ws_c.cell(row=6, column=9, value="Gastos Individual  ").font = Font(bold=True)
    encab = {
        1: "Dia ", 2: "Fecha ", 3: "Venta diaria ", 4: "Transferencia jesus",
        5: "Transferencia ", 6: "Salarios ", 7: "Carros  Corriente y Otras ",
        8: "Pagos  de la Caja ", 9: "Jesus ", 10: "Enrique ", 11: "Faltante",
        12: "Perdida de ganancia  por venta al COSTO ",
        13: "Ganancias por Venta (Elevacion de Precios) ", 14: "Sobrante",
        15: "EFECTIVO",
    }
    for c_idx, txt in encab.items():
        cell = ws_c.cell(row=7, column=c_idx, value=txt)
        cell.font = Font(bold=True); cell.fill = header_fill; cell.border = thin_border

    rng = (dias[0], dias[-1] + " 23:59:59")
    with get_conn() as conn:
        venta_d = _serie_por_dia(conn,
            """SELECT substr(fecha,1,10) d, COALESCE(SUM(total),0)
               FROM ventas WHERE estado!='cancelada' AND fecha>=? AND fecha<=?
               GROUP BY substr(fecha,1,10)""", rng)
        transf_d = _serie_por_dia(conn,
            """SELECT substr(fecha,1,10) d, COALESCE(SUM(subtotal_transferencia),0)
               FROM ventas WHERE estado!='cancelada' AND fecha>=? AND fecha<=?
               GROUP BY substr(fecha,1,10)""", rng)
        perd_d = _serie_por_dia(conn,
            """SELECT substr(v.fecha,1,10) d, COALESCE(SUM(vd.perdida_ganancia),0)
               FROM venta_detalle vd JOIN ventas v ON vd.venta_id=v.id
               WHERE v.estado!='cancelada' AND v.fecha>=? AND v.fecha<=?
               GROUP BY substr(v.fecha,1,10)""", rng)
        elev_d = _serie_por_dia(conn,
            """SELECT substr(v.fecha,1,10) d, COALESCE(SUM(vd.ganancia_elevacion),0)
               FROM venta_detalle vd JOIN ventas v ON vd.venta_id=v.id
               WHERE v.estado!='cancelada' AND v.fecha>=? AND v.fecha<=?
               GROUP BY substr(v.fecha,1,10)""", rng)

    fila = 10
    for dia in dias:
        d = datetime.strptime(dia, "%Y-%m-%d")
        ws_c.cell(row=fila, column=1, value=_dia_nombre(dia))
        fcell = ws_c.cell(row=fila, column=2, value=d)
        fcell.number_format = "dd/mm/yy"
        ws_c.cell(row=fila, column=3, value=venta_d.get(dia, 0) or 0).number_format = money_fmt
        ws_c.cell(row=fila, column=5, value=transf_d.get(dia, 0) or 0).number_format = money_fmt
        ws_c.cell(row=fila, column=12, value=perd_d.get(dia, 0) or 0).number_format = money_fmt
        ws_c.cell(row=fila, column=13, value=elev_d.get(dia, 0) or 0).number_format = money_fmt
        ef = ws_c.cell(row=fila, column=15,
                       value=(f"=C{fila}-D{fila}-E{fila}-F{fila}-G{fila}-H{fila}"
                              f"-I{fila}-J{fila}-K{fila}-L{fila}+M{fila}"))
        ef.number_format = money_fmt
        fila += 1

    tot = fila
    ws_c.cell(row=tot, column=1, value="TOTAL ").font = Font(bold=True)
    ws_c.cell(row=tot, column=2, value="Cierre").font = Font(bold=True)
    for c_idx in range(3, 16):
        letra = gl(c_idx)
        tc = ws_c.cell(row=tot, column=c_idx, value=f"=SUM({letra}10:{letra}{tot - 1})")
        tc.number_format = money_fmt
        tc.font = Font(bold=True)

    # Resumen P&L semanal (valores calculados por compute_cuadre, global)
    cuadre = compute_cuadre(dias[0], dias[-1] + " 23:59:59", 2, 20.0, 0.0)
    g = cuadre["gastos"]
    resumen = [
        ("Total de Venta Semanal por Sistema", cuadre["venta_total"]),
        ("Perdida de ganancia por venta al COSTO", cuadre["perdida_ganancia"]),
        ("Venta Real", cuadre["venta_real"]),
        ("Total de Venta al Precio de Costo", cuadre["venta_costo"]),
        ("Utilidad Bruta", cuadre["utilidad_bruta"]),
        ("Gastos Semanal (Salarios + Transporte)", g["salarios"] + g["transporte"]),
        ("ONAT, Arrendamiento y Contador", g["onat"] + g["arrendamiento"] + g["contador"]),
        ("Pago de Estimulacion", g["estimulacion"]),
        ("Utilidad Neta", cuadre["utilidad_neta"]),
        (f"Reserva ({cuadre['reserva_pct']}%)", cuadre["reserva"]),
        ("Dividendos a Repartir", cuadre["dividendos"]),
        (f"Por Socio ({cuadre['socios']} socios)", cuadre["por_socio"]),
        ("", None),
        ("Ganancia por Elevacion de Precios", cuadre["ganancia_elevacion"]),
        ("Venta x libreta (credito, por cobrar)", cuadre["venta_libreta"]),
        ("Cobros de libreta en el periodo", cuadre["cobros_libreta"]),
        ("Bajas / mermas (al costo)", cuadre["bajas_total"]),
    ]
    # Desglose de transferencia por persona/socio
    tps = cuadre.get("transferencia_por_socio") or {}
    if tps:
        resumen.append(("", None))
        resumen.append(("Transferencia por persona:", None))
        for soc, monto in tps.items():
            resumen.append((f"   {soc}", monto))
    rr = tot + 2
    for label, val in resumen:
        ws_c.cell(row=rr, column=1, value=label).font = Font(bold=True)
        if isinstance(val, (int, float)):
            vc = ws_c.cell(row=rr, column=5, value=val)
            vc.number_format = money_fmt
        rr += 1

    # ── Sub-tablas: entradas/efectivo por categoria y salarios por persona ──
    rango = (dias[0], dias[-1] + " 23:59:59")
    with get_conn() as conn:
        entradas_cat = {}
        for nombre, m in conn.execute(
            """SELECT cat.nombre, COALESCE(SUM(cd.subtotal),0) m
               FROM compra_detalle cd JOIN compras co ON cd.compra_id=co.id
               JOIN productos p ON cd.producto_id=p.id
               JOIN categorias cat ON p.categoria_id=cat.id
               WHERE co.fecha>=? AND co.fecha<=? GROUP BY cat.id""", rango).fetchall():
            entradas_cat[nombre] = entradas_cat.get(nombre, 0) + m
        for nombre, m in conn.execute(
            """SELECT cat.nombre, COALESCE(SUM(cd.cantidad_entregada*cd.costo_acordado),0) m
               FROM consignacion_detalle cd JOIN consignaciones co ON cd.consignacion_id=co.id
               JOIN productos p ON cd.producto_id=p.id
               JOIN categorias cat ON p.categoria_id=cat.id
               WHERE co.fecha_inicio>=? AND co.fecha_inicio<=? GROUP BY cat.id""", rango).fetchall():
            entradas_cat[nombre] = entradas_cat.get(nombre, 0) + m
        efectivo_cat = conn.execute(
            """SELECT cat.nombre, COALESCE(SUM(vd.subtotal),0) m
               FROM venta_detalle vd JOIN ventas v ON vd.venta_id=v.id
               JOIN productos p ON vd.producto_id=p.id
               JOIN categorias cat ON p.categoria_id=cat.id
               WHERE v.estado!='cancelada' AND v.fecha>=? AND v.fecha<=?
               GROUP BY cat.id ORDER BY m DESC""", rango).fetchall()
        salarios = conn.execute(
            """SELECT COALESCE(NULLIF(socio,''),'(sin socio)') s, COALESCE(SUM(monto),0) m
               FROM gastos WHERE tipo='salario' AND fecha>=? AND fecha<=?
               GROUP BY COALESCE(NULLIF(socio,''),'(sin socio)')""", rango).fetchall()

    def _tabla(titulo, filas):
        nonlocal rr
        rr += 2
        ws_c.cell(row=rr, column=1, value=titulo).font = header_font
        tot_t = 0
        for nombre, val in filas:
            rr += 1
            ws_c.cell(row=rr, column=1, value=nombre)
            ws_c.cell(row=rr, column=5, value=val).number_format = money_fmt
            tot_t += val or 0
        rr += 1
        ws_c.cell(row=rr, column=1, value="TOTAL").font = Font(bold=True)
        tc = ws_c.cell(row=rr, column=5, value=tot_t)
        tc.number_format = money_fmt
        tc.font = Font(bold=True)

    _tabla("Desglose de Entradas del Bazar (por categoria)",
           sorted(entradas_cat.items(), key=lambda x: -x[1]))
    _tabla("Efectivo del Bazar (por categoria)",
           [(r[0], r[1]) for r in efectivo_cat])
    _tabla("Salarios (por persona)", [(r[0], r[1]) for r in salarios])

    ws_c.column_dimensions["A"].width = 38
    for c_idx in range(2, 16):
        ws_c.column_dimensions[gl(c_idx)].width = 13

    # ── Hojas por vendedor / consignador (estructura simplificada del Grupo 3) ──
    # No | Costo | Producto | Inicio | Entradas | Deudas | Bajas | Ventas |
    # Importe al costo | Final
    rango_v = (dias[0], dias[-1] + " 23:59:59")
    with get_conn() as conn:
        vendedores = [r[0] for r in conn.execute(
            "SELECT DISTINCT consignador FROM productos "
            "WHERE activa=1 AND consignador IS NOT NULL AND consignador!='' "
            "ORDER BY consignador").fetchall()]
        # Series por producto (en el rango) reutilizadas para todos los vendedores
        ent_compra = _serie_por_dia(conn,
            """SELECT cd.producto_id, COALESCE(SUM(cd.cantidad),0)
               FROM compra_detalle cd JOIN compras c ON cd.compra_id=c.id
               WHERE c.fecha>=? AND c.fecha<=? GROUP BY cd.producto_id""", rango_v)
        ent_consig = _serie_por_dia(conn,
            """SELECT cd.producto_id, COALESCE(SUM(cd.cantidad_entregada),0)
               FROM consignacion_detalle cd JOIN consignaciones c ON cd.consignacion_id=c.id
               WHERE c.fecha_inicio>=? AND c.fecha_inicio<=? GROUP BY cd.producto_id""", rango_v)
        ven_prod = _serie_por_dia(conn,
            """SELECT vd.producto_id, COALESCE(SUM(vd.cantidad),0)
               FROM venta_detalle vd JOIN ventas v ON vd.venta_id=v.id
               WHERE v.estado!='cancelada' AND v.fecha>=? AND v.fecha<=?
               GROUP BY vd.producto_id""", rango_v)
        deu_prod = _serie_por_dia(conn,
            """SELECT cd.producto_id, COALESCE(SUM(cd.cantidad),0)
               FROM credito_detalle cd JOIN creditos c ON cd.credito_id=c.id
               WHERE c.fecha>=? AND c.fecha<=? GROUP BY cd.producto_id""", rango_v)
        baj_prod = _serie_por_dia(conn,
            """SELECT producto_id, COALESCE(SUM(cantidad),0) FROM bajas
               WHERE fecha>=? AND fecha<=? GROUP BY producto_id""", rango_v)
        snap_prod = {}
        for sr in conn.execute(
            "SELECT producto_id, stock FROM stock_snapshots WHERE fecha<=? "
            "ORDER BY producto_id, fecha, id", (dias[0] + " 23:59:59",)).fetchall():
            snap_prod[sr[0]] = sr[1]

        for vend in vendedores:
            prods_v = conn.execute(
                "SELECT id, nombre, costo, stock_inicial, stock_actual FROM productos "
                "WHERE activa=1 AND consignador=? ORDER BY nombre", (vend,)).fetchall()
            safe_title = vend.translate(str.maketrans({c: "-" for c in "\\/?*[]:"}))[:31]
            ws_v = wb.create_sheet(title=safe_title)
            ws_v["A1"] = "Basar Ojos Color Sol"
            ws_v["A1"].font = title_font
            ws_v["A2"] = f"{vend} · {titulo_semana}"
            enc_v = {1: "No", 2: "Costo", 3: "Producto", 4: "Inicio", 5: "Entradas",
                     6: "Deudas", 7: "Bajas", 8: "Ventas", 9: "Importe al costo",
                     10: "Final"}
            for c_idx, txt in enc_v.items():
                hc = ws_v.cell(row=4, column=c_idx, value=txt)
                hc.font = Font(bold=True); hc.fill = header_fill; hc.border = thin_border
            rv = 4
            for i, p in enumerate(prods_v, 1):
                rv += 1
                pid = p["id"]
                inicio = snap_prod.get(pid, p["stock_inicial"] or 0)
                entradas = (ent_compra.get(pid, 0) or 0) + (ent_consig.get(pid, 0) or 0)
                deudas = deu_prod.get(pid, 0) or 0
                bajas_u = baj_prod.get(pid, 0) or 0
                ventas_u = ven_prod.get(pid, 0) or 0
                ws_v.cell(row=rv, column=1, value=i)
                cb = ws_v.cell(row=rv, column=2, value=p["costo"]); cb.number_format = money_fmt
                ws_v.cell(row=rv, column=3, value=p["nombre"])
                ws_v.cell(row=rv, column=4, value=inicio)
                ws_v.cell(row=rv, column=5, value=entradas)
                ws_v.cell(row=rv, column=6, value=deudas)
                ws_v.cell(row=rv, column=7, value=bajas_u)
                ws_v.cell(row=rv, column=8, value=ventas_u)
                ws_v.cell(row=rv, column=9, value=f"=H{rv}*B{rv}").number_format = money_fmt
                # Final = Inicio + Entradas - Ventas - Deudas - Bajas
                ws_v.cell(row=rv, column=10, value=f"=D{rv}+E{rv}-H{rv}-F{rv}-G{rv}")
            ws_v.column_dimensions["C"].width = 28
            for col in ("A", "B", "D", "E", "F", "G", "H", "I", "J"):
                ws_v.column_dimensions[col].width = 11

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"semana_{desde}_{hasta}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
